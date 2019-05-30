from django.http import HttpResponse
from django.shortcuts import render, redirect
from xlrd import open_workbook
import xlrd
from openpyxl import Workbook
import pandas as pd
import openpyxl
import os
import xlwt

def data(request):
	workbook = open_workbook('Corp+Inc._Liste+de+comptes.xlsx')
	workbook = open_workbook('Corp+Inc._Liste+de+comptes.xlsx', on_demand = True)
	worksheet = workbook.sheet_by_name('Liste de comptes')
	sheet = workbook.sheet_by_index(0)

	rows = []
	c = 0
	for i in range(4, sheet.nrows):
	    columns = []

	    for j in range(sheet.ncols):
	        columns.append(sheet.cell(i, j).value)

	        data = sheet.cell(i, j).value
	        c += 1

	        if data == "Sous-type":
	        	dataSous = sheet.cell(i, c).value
	        	print(dataSous)
	        	col_name = data

        		createFolder(col_name)


	return HttpResponse(col_name)


def createFolder(directory):
	try:
		if not os.path.exists(directory):
			os.makedirs(directory)
	except OSError:
		print('Error: Creating directory. ' + directory)


def createSubFolder(directory):
	try:
		if not os.path.exists(directory):
			os.makedirs(directory)
	except OSError:
		print('Error: Creating directory. ' + directory)


def createFile(file):
	try:
		if not os.path.exists(file):
			os.startfile(file)
	except OSError:
		print('Error: Creating file. ' + file)